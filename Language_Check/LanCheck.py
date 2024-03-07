import os, openpyxl,argparse
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, colors
from openpyxl.comments import Comment
from langdetect import detect
from langdetect import detect_langs
from langdetect import DetectorFactory


# 从xlutils模块中导入copy这个函数

def updateExcel(file_path,x,y):
    '''''
    检查多语言文档中的文案格式是否正确
    :param file_path:传入一个excel文件，或者文件的绝对路径
    :param x：传入一个开始执行的行数
    :param y：传入一个开始执行的列数
    '''
    try:
        book = openpyxl.load_workbook(file_path)
    except Exception as e:
        # 如果路径不在或者excel不正确，返回报错信息
        print('路径不在或者excel不正确', e)
        return e
    else:
        sheet = book.active  # 取第一个sheet页
        rows_num = sheet.max_row + 1  # 取这个sheet页的所有行数
        clos_num = sheet.max_column + 1
        for r in list(range(rows_num))[x:]:  # 取第4行到最后一行
            for c in list(range(clos_num))[y:]:  # 取第9列到最后一行
                b = sheet.cell(row=r, column=c)
                a = b.value
                b.fill = None #清空Excel背景色
                b.comment = None #清空Excel备注

                ##检查是否有内容为空###
                if a == None:
                    commentexcel(b, r, c, '内容为空')
                    continue

                # ###检查内容是否超出字符数###
                # length = sheet.cell(row=r, column=8).value
                # try:
                #     float(length)
                # except:
                #     print("跳过" + str(r) + '行' + str(c) + '列')
                # else:
                #     if len(str(a)) > int(length):
                #         commentexcel(b, r, c, '超出预定字符')

                # identifier = LanguageIdentifier.from_modelstring(model, norm_probs=True)
                # language, confidence = identifier.classify(a)
                # lan = sheet.cell(row=3, column=c).value
                # print(lan)
                # if language != lan and confidence == 1:
                #     commentexcel(b, r, c, '该文案不是' + lan + '语言,是' + language + '语言' + str(confidence))

                ###检查是否有多余的空格###
                word = a.split(' ')
                for i in list(range(len(word))):
                    if word[i] == "":
                        commentexcel(b, r, c, '有多余空格')
                        break

                ###检查%的输入法问题：不能出现中文格式的%
                if a.find("％") != -1:
                    commentexcel(b, r, c, '"%"格式为中文')

                ###检查%S的大小写问题：不能出现大写的%S
                if a.find("%S") != -1:
                    commentexcel(b, r, c, '"%S"大小写错误')
                elif a.find("% S") != -1:
                    commentexcel(b, r, c, '"%S"大小写错误')

                ###检查%前后空格：%前后不能都为空格
                if a.find("%") != -1:
                    number = a.index("%")
                    if number < len(a) - 1:
                        if a[number - 1] == " " and a[number + 1] == " ":
                            commentexcel(b, r, c, '"%"前后都为空')
                    elif number == len(a) - 1 and number != 0:
                        if a[number - 1] == " ":
                            commentexcel(b, r, c, '"%"前后都为空')
                    elif number == 0 and len(a) != 1:
                        if a[number + 1] == " ":
                            commentexcel(b, r, c, '"%"前后都为空')
                    elif number == 0 and len(a) == 1:
                        commentexcel(b, r, c, '"%"前后都为空')

                ###检查"/n"前后是否有空格：/n前后不能都为空格
                if a.find("/n") != -1:
                    if a.find(" /n") != -1 or a.find("/n ") != -1:
                        commentexcel(b, r, c, '"/n"前后有空格')

                ###检查"$"后是否有空格：$前后不能都为空格
                if a.find("$") != -1:
                    if a.find(" $s") != -1:
                        commentexcel(b, r, c, '"$"后有空格/"$s"前边有空格')
                    elif a.find("$ ") != -1:
                        commentexcel(b, r, c, '"$"后有空格/"$s"前边有空格')
                        commentexcel(b, r, c, '"$"后有空格/"$s"前边有空格')
                    elif a.find("% ") != -1:
                        commentexcel(b, r, c, '"%"后有空格')

                ###检查非中文语系是否有中文标点符号
                language = sheet.cell(row=3, column=c).value
                if language != "zh-CN" and language != "zh-TW" and language != "ja":
                    if a.find("！") != -1:
                        commentexcel(b, r, c, '"！"是中文标点')
                    elif a.find("。") != -1:
                        commentexcel(b, r, c, '"。"是中文标点')
                    elif a.find("，") != -1:
                        commentexcel(b, r, c, '"，"是中文标点')
                    elif a.find("？") != -1:
                        commentexcel(b, r, c, '"？"是中文标点')
                    elif a.find("：") != -1:
                        commentexcel(b, r, c, '"："是中文标点')
                    elif a.find("（") != -1:
                        commentexcel(b, r, c, '"（"是中文标点')
                    elif a.find("）") != -1:
                        commentexcel(b, r, c, '"）"是中文标点')
                    # elif a.find("'") != -1:
                    #     commentexcel(b, r, c, '"""是中文标点')
                    # elif a.find("'") != -1:
                    #     commentexcel(b, r, c, '"""是中文标点')
                else: ###检查中文语系是否有英文标点符号
                    if a.find("!") != -1:
                        commentexcel(b, r, c, '"!"是英文标点')
                    elif a.find(",") != -1:
                        commentexcel(b, r, c, '","是英文标点')
                    elif a.find("?") != -1:
                        commentexcel(b, r, c, '"?"是英文标点')
                    elif a.find(":") != -1:
                        commentexcel(b, r, c, '":"是英文标点')
                    elif a.find("(") != -1:
                        commentexcel(b, r, c, '"("是英文标点')
                    elif a.find(")") != -1:
                        commentexcel(b, r, c, '")"是英文标点')
                    elif a.find('"') != -1:
                        commentexcel(b, r, c, '"是英文标点')

                ###检查特殊符号是否遗漏
                d = sheet.cell(row=r, column=9).value
                if d != None:
                    if d.find("%") != -1:
                        if a.find("%") == -1:
                            commentexcel(b, r, c, '缺少特殊字符"%"')
                if d.find("%s") != -1:
                    if a.find("%s") == -1:
                        commentexcel(b, r, c, '缺少特殊字符"%s"')
                if d.find("#") != -1:
                    if a.find("#") == -1:
                        commentexcel(b, r, c, '缺少特殊字符"#"')
                if d.find("$") != -1:
                    if a.find("$") == -1:
                        commentexcel(b, r, c, '缺少特殊字符"$"')
                if d.find("$s") != -1:
                    if a.find("$s") == -1:
                        commentexcel(b, r, c, '缺少特殊字符"$s"')
                if d.find("/n") != -1:
                    if a.find("/n") == -1:
                        commentexcel(b, r, c, '缺少特殊字符"/n"')
                else:
                    continue
        book.save(file_path)

def CompairText(file_path,x,y):
    '''''
    检查多语言文档中的文案是否匹配
    :param file_path:传入一个excel文件，或者文件的绝对路径
    :param x：传入一个开始执行的行数
    :param y：传入一个开始执行的列数
    '''
    try:
        book = openpyxl.load_workbook(file_path)
    except Exception as e:
        # 如果路径不在或者excel不正确，返回报错信息
        print('路径不在或者excel不正确', e)
        return e
    else:
        sheet = book.active  # 取第一个sheet页
        clos_num = sheet.max_column + 1 # 取这个sheet页的所有列数
        for c in list(range(clos_num))[y:]:  # 取第9列到最后一行
            b16 = sheet.cell(row=7, column=c)
            b17 = sheet.cell(row=9, column=c)
            b18 = sheet.cell(row=10, column=c)

            b18.fill = None  # 清空Excel背景色
            b18.comment = None  # 清空Excel备注

            if b17.value not in b16.value:
                commentexcel(b17, 9, c, '内容不匹配')
            if b18.value not in b16.value:
                commentexcel(b18, 10, c, '内容不匹配')
            else:
                continue
        book.save(file_path)

def commentexcel(b, r, c, string):
    # fill_3 = PatternFill("solid", fgColor="ffc7ce")  # 红色
    # b.fill = fill_3
    #
    comment = Comment(string, "tester")
    b.comment = comment
    print(str(r) + '行' + str(c) + '列' + string)
    return

def CompairLan(file_path,x,y):
    try:
        book = openpyxl.load_workbook(file_path)
    except Exception as e:
        # 如果路径不在或者excel不正确，返回报错信息
        print('路径不在或者excel不正确', e)
        return e
    else:
        sheet = book.active  # 取第一个sheet页
        rows_num = sheet.max_row + 1  # 取这个sheet页的所有行数
        clos_num = sheet.max_column + 1
        for r in list(range(rows_num))[x:]:  # 取第4行到最后一行
            for c in list(range(clos_num))[y:]:  # 取第9列到最后一行
                b = sheet.cell(row=r, column=c)
                a = b.value
                key = sheet.cell(row=2,column=c)
                b.fill = None #清空Excel背景色
                b.comment = None #清空Excel备注
                DetectorFactory.seed = 0
                if detect(a) not in key.value:
                    print(key.value)
                    print(detect(a))
                    commentexcel(b, r, c, "不是这个国家的语言")
        book.save(file_path)

###直接在代码中修改执行文件路径
# # 获取项目路径
# project_path = os.path.abspath(os.path.join(os.path.dirname(os.path.split(os.path.realpath(__file__))[0]), '.'))
#
# # 测试表格存放路径
# test_cases_path = project_path + "/Language_Check/"
#
# # 执行
# updateExcel(test_cases_path + "pbn的副本.xlsx")
parser = argparse.ArgumentParser(description='检查多语言文档格式是否正确')
parser.add_argument('file', type=str, help='检测文件路径')
parser.add_argument('row', type=int, help='从第x行开始检测')
parser.add_argument('clos', type=int, help='从第y列开始检查')
args = parser.parse_args()

# 获取文件路径参数
file_path = args.file
x = args.row
y = args.clos

# 检查文件格式
# updateExcel(file_path,x,y)
# CompairText(file_path,x,y)
CompairLan(file_path, x, y)
